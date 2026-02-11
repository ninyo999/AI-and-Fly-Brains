import customtkinter as ctk
import openpyxl
from pathlib import Path
from arduino_ctrl import ArduinoHandler

class BaseModal(ctk.CTkToplevel):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.attributes("-topmost", True)
        self.overrideredirect(True)
        self.wait_visibility()
        self.grab_set()  

    def close(self):
        self.grab_release()
        self.destroy()

class AddExpModal(BaseModal):
    def __init__(self, parent, save_callback, experiment_path_root):
        super().__init__(parent)
        self.arduino = ArduinoHandler()
        self.save_callback = save_callback
        self.exp_path_root = Path(experiment_path_root)
        
        # UI Setup
        self.geometry("900x700+230+50")
        self.configure(fg_color="#E5E5E5", corner_radius=30)
        
        # Header Section
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=40, pady=20)
        ctk.CTkLabel(header, text="Add New Experiment", text_color="#B32442", 
                     font=ctk.CTkFont(size=32, weight="bold")).pack(side="left")
        ctk.CTkButton(header, text="X", width=40, height=40, corner_radius=20, 
                      fg_color="transparent", text_color="#B32442", border_width=2,
                      border_color="#B32442", command=self.close).pack(side="right")

        # Experiment Info Inputs
        self.title_e = self.create_input("Title:")
        self.drug_e = self.create_input("Drug use:")
        self.remark_e = self.create_textbox("Remark:")

        # Dark field LEDs Section
        self.create_section("Dark field LEDs")
        df_frame = ctk.CTkFrame(self, fg_color="transparent")
        df_frame.pack(fill="x", padx=60)
        self.d_duty = self.create_grid_input(df_frame, "Duty Cycle(%)", 0, "60")
        self.d_freq = self.create_grid_input(df_frame, "Frequency(Hz)", 1, "500")
        self.d_time = self.create_grid_input(df_frame, "Active time(s)", 2, "40")

        # Optogenetic LEDs Section
        self.create_section("Optogenetic LEDs")
        op_frame = ctk.CTkFrame(self, fg_color="transparent")
        op_frame.pack(fill="x", padx=60)
        self.o_duty = self.create_grid_input(op_frame, "Duty Cycle(%)", 0, "100")
        self.o_freq = self.create_grid_input(op_frame, "Frequency(Hz)", 1, "500")
        self.o_len = self.create_grid_input(op_frame, "Flash length(s)", 2, "5")
        self.o_delay = self.create_grid_input(op_frame, "Initial delay(s)", 3, "8")

        # Start Button
        ctk.CTkButton(self, text="Start Experiment", fg_color="#B32442", height=50, 
                      width=300, corner_radius=15, command=self.handle_start).pack(pady=30)

    def save_to_excel(self, folder_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Experiment Details"
        title_val = self.title_e.get()

        template_data = [
            ("Title:", title_val),
            ("Drug use:", self.drug_e.get()),
            ("Remark:", self.remark_e.get("1.0", "end-1c")),
            ("", ""), # Row 4
            ("Dark field LEDs", ""),
            ("duty cycle:", self.d_duty.get()),
            ("frequency:", self.d_freq.get()),
            ("active time:", self.d_time.get()),
            ("", ""), # Row 9
            ("Optogenetic LEDs", ""),
            ("duty cycle:", self.o_duty.get()),
            ("frequency:", self.o_freq.get()),
            ("flash length:", self.o_len.get()),
            ("initial delay:", self.o_delay.get())
        ]

        for r_idx, (label, val) in enumerate(template_data, start=1):
            ws.cell(row=r_idx, column=1, value=label)
            ws.cell(row=r_idx, column=2, value=val)

        filename = f"experiment_data_{title_val}.xlsx"
        wb.save(folder_path / filename)

    def handle_start(self):
        title = self.title_e.get()
        if title:
            # Create Folder
            new_dir = self.exp_path_root / title
            new_dir.mkdir(exist_ok=True)
            
            # Save Excel Template
            self.save_to_excel(new_dir)
            
            # Send Data to Arduino Handler
            led_data = {
                'dark_duty': self.d_duty.get(), 'dark_freq': self.d_freq.get(), 'dark_time': self.d_time.get(),
                'opto_duty': self.o_duty.get(), 'opto_freq': self.o_freq.get(), 'opto_len': self.o_len.get(), 
                'opto_delay': self.o_delay.get()
            }
            self.arduino.send_led_data(led_data)

            self.save_callback() 
            self.close()

    def create_input(self, text):
        ctk.CTkLabel(self, text=text, text_color="#B32442", font=("Arial", 14, "bold")).pack(anchor="w", padx=60)
        e = ctk.CTkEntry(self, width=780, fg_color="white", border_width=0, text_color="black")
        e.pack(pady=(2, 10))
        return e

    def create_textbox(self, text):
        ctk.CTkLabel(self, text=text, text_color="#B32442", font=("Arial", 14, "bold")).pack(anchor="w", padx=60)
        t = ctk.CTkTextbox(self, width=780, height=100, fg_color="white", text_color="black")
        t.pack(pady=5)
        return t

    def create_section(self, text):
        ctk.CTkLabel(self, text=text, text_color="#B32442", font=("Arial", 20, "bold")).pack(anchor="w", padx=60, pady=10)

    def create_grid_input(self, master, label, col, default_val=""):
        f = ctk.CTkFrame(master, fg_color="transparent")
        f.grid(row=0, column=col, padx=(0, 20))
        ctk.CTkLabel(f, text=label, text_color="#B32442").pack(side="left")
        e = ctk.CTkEntry(f, width=80, fg_color="white", border_width=0, text_color="black")
        e.insert(0, default_val)
        e.pack(side="left", padx=5)
        return e